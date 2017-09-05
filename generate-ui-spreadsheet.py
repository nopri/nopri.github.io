#!/usr/bin/env python
#
#
# generate-ui-spreadsheet.py
#
# (c) Noprianto <nop@noprianto.com>
# 2017
# License: GPL
# https://github.com/nopri
#
# Based on some codes from https://github.com/nopri/sqliteboy
#


import sys

try:
    import os
    import time

    import openpyxl
    from openpyxl.utils import _get_column_letter as column_letter
except Exception, e:
    sys.exit(e)


NAME = 'generate-ui-spreadsheet'
VERSION = '0.4'
COPYRIGHT = '(c) Noprianto <nop@noprianto.com>, 2017'
TITLE = '%s version %s, %s' %(NAME, VERSION, COPYRIGHT)
CODE_OK = 0
CODE_ERROR = 1
CODE_ERROR_TYPE = 2
CODE_ERROR_FILE = 3
CODE_ERROR_EXISTS = 4
CODE_ERROR_FUNCTION = 5
CODE_ERROR_UI = 6
CODE_ERROR_CODE = 7
CODE_ERROR_OUTPUT = 8
TYPE = {
        'html':  ('.html', 'generate_html'),
        'python-tk': ('.py', 'generate_python_tk'),
    }
TYPE_KEY = TYPE.keys()
TYPE_STR = ' '.join(TYPE_KEY)
WIDGET_SEP = ':'
WIDGET_TYPE_TEXT = 'text'
WIDGET_TYPE_CHECK = 'check'
WIDGET_TYPE_BUTTON = 'button'
WIDGET_TYPE_COMBO = 'combo'
WIDGETS = [
        WIDGET_TYPE_BUTTON,
        WIDGET_TYPE_CHECK,
        WIDGET_TYPE_COMBO,
        WIDGET_TYPE_TEXT,
    ]
WIDGET_COMBO_SOURCE = 'source'
WIDGET_COMBO_DATA = 'data'
WIDGET_TEXT_SIZE = 'size'
WIDGET_TEXT_MAXLENGTH = 'maxlength'


symbol_html = []
symbol_python_tk = []


def strs(s):
    return str(s)


def is_integer(n):
    if isinstance(n, (int, long)):
        return True
    #
    return False


def is_float(n):
    if isinstance(n, float):
        return True
    #
    return False


def log(msg, newline=1, stream=sys.stdout):
    try:
        newline = int(newline)
    except ValueError:
        newline = 0
    #
    end = os.linesep * newline
    #
    if not stream in [sys.stdout, sys.stderr]:
        return
    #
    stream.write('%s%s' %(msg, end) )


def log_err(msg, newline=1):
    return log(msg, newline, stream=sys.stderr)


def get_ui_sheet(book):
    try:
        return book.worksheets[0]
    except:
        pass
    return None


def get_code_sheet(book):
    try:
        return book.worksheets[1]
    except:
        pass
    return None


def get_ui_size(sheet):
    ret = (-1, -1)
    try:
        ret = (sheet.max_column, sheet.max_row)
    except:
        pass
    return ret


def get_cell(column, row):
    return '%s%s' %(column_letter(column), row)


def get_code_properties(value, sheet):
    ret = ''
    #
    if not len(value) > 2:
        return ret
    #
    max_row = sheet.max_row
    column = 1
    column_properties_start = 2
    start = 1
    #
    for i in range(1, max_row + 1):
        cell = sheet.cell(get_cell(column, i))
        if cell.value == value[2]:
            start = i + 1
            break
    #
    ret = {}
    while True:
        cell = sheet.cell(get_cell(column, start))
        cell2 = sheet.cell(get_cell(column_properties_start, start))
        if not cell.value:
            prop = cell2.value
            if prop:
                prop = strs(prop).lower()
                data = []
                col = column_properties_start + 1
                while True:
                    cell_prop = sheet.cell(get_cell(col, start))
                    cell_prop_value = cell_prop.value
                    if cell_prop_value:
                        if cell_prop.data_type == 'n':
                            if is_integer(cell_prop_value):
                                cell_prop_value = int(cell_prop_value)
                            elif is_float(cell_prop_value):
                                cell_prop_value = float(cell_prop_value)
                        data.append(cell_prop_value)
                    else:
                        break
                    col += 1
                ret[prop] = data
            start += 1
        #
        cell3 = sheet.cell(get_cell(column_properties_start, start))
        if not cell3.value:
            break
    #
    return ret


def get_widget_type(value):
    value = strs(value)
    ret = ['', value, value]
    #
    if WIDGET_SEP in value:
        temp = value.split(WIDGET_SEP)
        if temp and len(temp) > 1:
            temp_type = temp[0].lower()
            temp_value = temp[1]
            if temp_type in WIDGETS:
                ret = [temp_type, temp_value, value]
    #
    return ret


def get_widget_combo_items(properties, single=False):
    ret = []
    source = properties.get(WIDGET_COMBO_SOURCE)
    data = properties.get(WIDGET_COMBO_DATA)
    if not source and data:
            for i in data:
                if single:
                    item = i
                else:
                    item = (i, i)
                ret.append(item)
    return ret


def get_widget_property(properties, key):
    ret = None
    temp = properties.get(key, [])
    if temp:
        ret = temp[0]
    return ret


def interpret_cell_html_color(rgb, style):
    ret = ''
    rgb = strs(rgb)
    rgb = rgb[2:]
    if rgb != '000000':
        ret = '%s: #%s;' %(style, rgb)
    return ret


def interpret_cell_python_tk_color(rgb, style):
    ret = ''
    rgb = strs(rgb)
    rgb = rgb[2:]
    if rgb != '000000':
        ret = "['%s'] = '#%s'" %(style, rgb)
    return ret


def interpret_cell_html_widget(widget, styles, sheet):
    ret = ''
    if isinstance(widget, (list, tuple)):
        if  len(widget) >= 2:
            widget_type = widget[0]
            widget_value = widget[1]
            prop = get_code_properties(widget, sheet)
            style = ' '.join(styles).strip()
            style_add = ''
            if style:
                style_add = ' style="%s"' %(style)
            if not widget_type:
                ret = '''
                    <span%s>%s</span>
                ''' %(style_add, widget_value)
            elif widget_type == WIDGET_TYPE_BUTTON:
                ret = '''
                    <input%s type='button' name='%s' value='%s'>
                ''' %(style_add, widget_value, widget_value)
            elif widget_type == WIDGET_TYPE_CHECK:
                ret = '''
                    <input%s type='checkbox' name='%s' value='%s'>%s
                ''' %(style_add, widget_value, widget_value, widget_value)
            elif widget_type == WIDGET_TYPE_COMBO:
                ret = '''
                    <select%s name='%s'>
                ''' %(style_add, widget_value)
                items = get_widget_combo_items(prop)
                for i in items:
                    ret += '''
                        <option value='%s'>%s</option>
                    ''' %(i[0], i[1])
                ret += '''
                    </select>
                '''
            elif widget_type == WIDGET_TYPE_TEXT:
                size_add = ''
                size = get_widget_property(prop, WIDGET_TEXT_SIZE)
                if size:
                    size_add = " size='%s' " %(size)
                maxlength_add = ''
                maxlength = get_widget_property(prop, WIDGET_TEXT_MAXLENGTH)
                if maxlength:
                    maxlength_add = " maxlength='%s' " %(size)
                ret = '''
                    <input%s type='text' name='%s'%s%s>
                ''' %(style_add, widget_value, size_add, maxlength_add)
    return ret


def get_var_python(value, default):
    num = len(symbol_python_tk)
    var = '%s_%s' %(default, num)
    ret = (var,)
    symbol_python_tk.append(ret)
    return ret


def get_var_python_tk(value, default='var'):
    return get_var_python(value, default)[0]


def interpret_cell_python_tk_widget(widget, styles, sheet, row, col):
    ret = ''
    if isinstance(widget, (list, tuple)):
        if  len(widget) >= 2:
            widget_type = widget[0]
            widget_value = widget[1]
            prop = get_code_properties(widget, sheet)
            var = get_var_python_tk(widget_value)
            if not widget_type:
                ret = """
        #%s: %s
        self.%s = Label()
        self.%s['text'] = '%s'
        self.%s.grid(column=%s, row=%s)
                """ %(widget_value, widget_type,
                    var, var, widget_value, var, col, row)
            elif widget_type == WIDGET_TYPE_BUTTON:
                ret = """
        #%s: %s
        self.%s = Button()
        self.%s['text'] = '%s'
        self.%s.grid(column=%s, row=%s)
                """ %(widget_value, widget_type,
                    var, var, widget_value, var, col, row)
            elif widget_type == WIDGET_TYPE_CHECK:
                ret = """
        #%s: %s
        self.%s = Checkbutton()
        self.%s['text'] = '%s'
        self.%s.grid(column=%s, row=%s)
                """ %(widget_value, widget_type,
                    var, var, widget_value, var, col, row)
            elif widget_type == WIDGET_TYPE_COMBO:
                items = get_widget_combo_items(prop, single=True)
                ret = """
        #%s: %s
        self.%s = ttk.Combobox(values=%s)
        self.%s.current(0)
        self.%s.grid(column=%s, row=%s)
                """ %(widget_value, widget_type, var, items, var,
                    var, col, row)
            elif widget_type == WIDGET_TYPE_TEXT:
                ret = """
        #%s: %s
        self.%s = Entry()
        self.%s.grid(column=%s, row=%s)
                """ %(widget_value, widget_type, var, var, col, row)

        #
        for i in styles:
            if i:
                ret += """
        self.%s%s
                """ %(var, i)
        #
    return ret


def interpret_cell_html(cell, sheet):
    ret = ''
    value = cell.value
    font_color = cell.font.color.rgb
    bg_color = cell.fill.fgColor.rgb
    if value:
        style_font_color = interpret_cell_html_color(
            font_color, 'color')
        style_bg_color = interpret_cell_html_color(
            bg_color, 'background-color')
        styles = [style_font_color, style_bg_color]
        ret = interpret_cell_html_widget(get_widget_type(value), styles,
            sheet)
    return ret


def interpret_cell_python_tk(cell, sheet, row, col):
    ret = ''
    value = cell.value
    font_color = cell.font.color.rgb
    bg_color = cell.fill.fgColor.rgb
    if value:
        style_font_color = interpret_cell_python_tk_color(
            font_color, 'fg')
        style_font_color_active = interpret_cell_python_tk_color(
            font_color, 'activeforeground')
        style_bg_color = interpret_cell_python_tk_color(
            bg_color, 'bg')
        style_bg_color_active = interpret_cell_python_tk_color(
            bg_color, 'activebackground')
        styles = [style_font_color, style_font_color_active,
                style_bg_color, style_bg_color_active
            ]
        ret = interpret_cell_python_tk_widget(get_widget_type(value), styles,
            sheet, row, col)
    return ret


def generate_html(book, output_file):
    max_column, max_row = get_ui_size(get_ui_sheet(book))
    column_width = int(round(100/max_column))
    sheet = get_ui_sheet(book)
    sheet_code = get_code_sheet(book)
    if max_column < 1 or max_row < 1 or sheet is None:
        return CODE_ERROR_UI
    #
    code = '''<!DOCTYPE html>
<!--%s-->
<!--%s-->
<html>
<head>
    <title>%s</title>
    <style>
    table
    {
        border-collapse : collapse;
        width           : 100%%;
    }
    td
    {
        padding         : 2px;
    }
    </style>
</head>
<body>
<table>
''' %(TITLE, time.asctime(), os.path.basename(output_file))
    #
    for row in range(1, max_row + 1):
        code += '''
    <tr>
    '''
        for col in range(1, max_column + 1):
            code += '''
        <td width='%s%%'>
        ''' %(column_width)
            cell = sheet.cell(get_cell(col, row))
            code += interpret_cell_html(cell, sheet_code)
            code += '''
        </td>
        '''
        code += '''
    </tr>
    '''
    #
    code += '''</table>
</body>
</html>
'''
    #
    try:
        with open(output_file, 'w') as f:
            f.write(code)
    except:
        return CODE_ERROR_OUTPUT
    return CODE_OK


def generate_python_tk(book, output_file):
    max_column, max_row = get_ui_size(get_ui_sheet(book))
    column_width = int(round(100/max_column))
    sheet = get_ui_sheet(book)
    sheet_code = get_code_sheet(book)
    if max_column < 1 or max_row < 1 or sheet is None:
        return CODE_ERROR_UI
    #
    code = '''
#
# %s
# %s

from Tkinter import *
import ttk

class Application(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.create_ui()

    def create_ui(self):
''' %(TITLE, time.asctime())

    #
    for row in range(1, max_row + 1):
        for col in range(1, max_column + 1):
            cell = sheet.cell(get_cell(col, row))
            code += interpret_cell_python_tk(cell, sheet_code, row, col)

    code += '''

def main():
    app = Application()
    app.master.title('%s')
    app.mainloop()


if __name__ == '__main__':
    main()

    ''' %(os.path.basename(output_file))
    #
    try:
        with open(output_file, 'w') as f:
            f.write(code)
    except:
        return CODE_ERROR_OUTPUT
    return CODE_OK


def generate(book, output_type, output_file):
    func_name = TYPE.get(output_type)[1]
    func = globals().get(func_name)
    if not func:
        log_err('Unknown function: %s' %(func_name))
        return CODE_ERROR_FUNCTION
    ret = func(book, output_file)
    return ret


def main(input_file, output_type):
    if not output_type in TYPE_KEY:
        log_err('Unknown type: %s' %(output_type))
        return CODE_ERROR_TYPE
    #
    try:
        book = openpyxl.load_workbook(input_file)
    except Exception, e:
        log_err(e)
        return CODE_ERROR_FILE
    #
    output_file = '%s%s' %(
        os.path.splitext(input_file)[0], TYPE.get(output_type)[0])
    if os.path.exists(output_file):
        log_err('Output file exists: %s' %(output_file))
        return CODE_ERROR_EXISTS
    #
    return generate(book, output_type, output_file)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        log(TITLE)
        log('%s <file> <type>' %(sys.argv[0]))
        log('\t <type>: %s' %(TYPE_STR))
        sys.exit(CODE_ERROR)
    else:
        input_file = os.path.abspath(sys.argv[1])
        output_type = sys.argv[2]
        ret = main(input_file, output_type)
        sys.exit(ret)
