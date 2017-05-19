#
# Simple Spreadsheet Test Case
# (c) Noprianto <nop@noprianto.com>
# GPL
#
#


import sys
import os
import time
from openpyxl import Workbook
from PySide import QtGui


APP_NAME = 'Simple Spreadsheet Test Case'
APP_VERSION = '0.1'
APP_TITLE = '%s %s' %(APP_NAME, APP_VERSION)
UI_WIDTH = 800
UI_HEIGHT = 600
DEF_ROWS = 100
DEF_COLS = 100
FORMAT_DATE_TIME = '%Y-%m-%d %H:%M:%S'
FILTER_FILE = '*.xlsx'
T_CELL_NAME = 'Cell Name'
T_CELL = 'Cell'
T_CASE = 'Case'
T_FORMULA = 'Formula'
T_ASSERT_RESULT_CELL = 'Assert Result On Cell'
T_RESULT_CELL = 'Result On Cell'
T_RESULT_EXPECTED = 'Expected Result'
T_GENERATE = 'Generate'
T_ERROR = 'Error'
T_INFO = 'Info'
T_PROVIDE_INPUT = 'Please provide valid input'
T_SAVE_TO = 'Saved to'
T_TEST_CASE = 'Test Case(s)'
T_RESULT = 'Result'
T_FAILED = 'Failed'
T_OK = 'OK'
T_TEST_START = 'Test Started On'
T_TEST_FINISH = 'Test Finished On'
T_OUTPUT_FILE = 'Output file name (*.xlsx)'


def try_number(n):
    res = n
    #
    try:
        float(n)
        is_number = True
    except:
        is_number = False
    #
    if is_number:
        if '.' in n:
            res = float(n)
        else:
            res = int(n)
    #
    return res
    

class Application(QtGui.QWidget):
    def __init__(self):
        super(Application, self).__init__()
        self.fixedHorizontalHeader = [T_CELL_NAME]
        self.fixedVerticalHeader = [T_RESULT_EXPECTED]
        self.createUI()
        
    def createTableHorizontalHeader(self):
        ret = self.fixedHorizontalHeader[:]
        counter = 1
        for i in range(DEF_COLS-len(ret)):
            temp = '%s %s' %(T_CASE, counter)
            ret.append(temp)
            counter += 1
        return ret  

    def createTableVerticalHeader(self):
        ret = self.fixedVerticalHeader[:]
        counter = 1
        for i in range(DEF_ROWS-len(ret)):
            temp = '%s %s' %(T_CELL, counter)
            ret.append(temp)
            counter += 1
        return ret          
        
    def checkCellFormat(self, text):
        return len(text) > 0
    
    def generate(self):
        time_start = time.strftime(FORMAT_DATE_TIME)
        #
        wb = Workbook()
        ws = wb.active
        ws.title = T_RESULT
        ws.cell(row=1, column=1).value = T_TEST_CASE
        #
        vfor = self.editFor.text().strip()
        vres = self.editRes.text().strip()
        vasr = self.editAsr.text().strip()
        if not self.checkCellFormat(vfor) or \
            not self.checkCellFormat(vres) or \
            not self.checkCellFormat(vasr):
                QtGui.QMessageBox.critical(self, T_ERROR, T_PROVIDE_INPUT)
                return
        #
        out_file = QtGui.QFileDialog.getSaveFileName(self, 
            caption=T_OUTPUT_FILE, 
            filter=FILTER_FILE)
        out_file = out_file[0].strip()
        if not out_file:
            return
        #
        startCol = 1 
        stopCol = DEF_COLS - len(self.fixedHorizontalHeader) 
        startRow = 0
        stopRow = DEF_ROWS - len(self.fixedVerticalHeader)
        numTest = 0
        for col in range(startCol, stopCol):
            cell = self.table.item(0, col) #expected result
            if not cell:
                break
            name = 'test_%s' %(col)
            sheet = wb.create_sheet(name)
            for row in range(startRow, stopRow):
                ncell = self.table.item(row, 0)
                if not ncell:
                    break
                cellName = ncell.text()
                if row == 0:
                    scell = sheet.cell(vasr)
                    scell.value = '=%s=%s' %(vres, cellName)
                if not self.checkCellFormat(cellName):
                    break
                cell = self.table.item(row, col)
                if not cell:
                    break
                scell = sheet.cell(cellName)
                scell.value = try_number(cell.text())
                scell = sheet.cell(vres)
                scell.value = vfor
            numTest += 1
            #
            ws.cell(row=numTest+1, column=1).value="[%s] %s" %(
                sheet.title, vasr)
            ws.cell(row=numTest+1, column=2).value="=%s!%s" %(
                sheet.title, vasr)
            #
        #
        ws.cell(row=numTest+3, column=1).value = T_FAILED
        ws.cell(row=numTest+3, column=2).value = '=CountIf(B2:B%s,0)' %(
            numTest+1)
        ws.cell(row=numTest+4, column=1).value = T_OK
        ws.cell(row=numTest+4, column=2).value = '=CountIf(B2:B%s,1)' %(
            numTest+1)
        ws.cell(row=numTest+6, column=1).value = T_TEST_START
        ws.cell(row=numTest+6, column=2).value = time_start
        ws.cell(row=numTest+7, column=1).value = T_TEST_FINISH
        ws.cell(row=numTest+7, column=2).value = time.strftime(FORMAT_DATE_TIME)
        #
        wb.save(out_file)
        msg = '%s %s (%s %s)' %(T_SAVE_TO, out_file, numTest, T_TEST_CASE)
        QtGui.QMessageBox.information(self, T_INFO, msg)                
        
    def createUI(self):
        grid = QtGui.QGridLayout()
        #
        lblFor = QtGui.QLabel(T_FORMULA)
        self.editFor = QtGui.QLineEdit()
        lblRes = QtGui.QLabel(T_RESULT_CELL)
        self.editRes = QtGui.QLineEdit()
        lblAsr = QtGui.QLabel(T_ASSERT_RESULT_CELL)
        self.editAsr = QtGui.QLineEdit()
        genBtn = QtGui.QPushButton(T_GENERATE)
        genBtn.clicked.connect(self.generate)
        #
        self.table = QtGui.QTableWidget(DEF_ROWS, DEF_COLS)
        self.table.setHorizontalHeaderLabels(self.createTableHorizontalHeader())
        self.table.setVerticalHeaderLabels(self.createTableVerticalHeader())        
        #
        grid.addWidget(lblFor, 0, 1, 1, 2)
        grid.addWidget(self.editFor, 0, 3, 1, 8)
        grid.addWidget(lblRes, 1, 1, 1, 2)
        grid.addWidget(self.editRes, 1, 3, 1, 8)
        grid.addWidget(lblAsr, 2, 1, 1, 2)
        grid.addWidget(self.editAsr, 2, 3, 1, 8)
        grid.addWidget(self.table, 3, 1, 10, 10)
        grid.addWidget(genBtn, 14, 1)        
        #
        self.setLayout(grid)
        self.resize(UI_WIDTH, UI_HEIGHT)
        self.setWindowTitle(APP_TITLE)
        self.show()


def main():
    app = QtGui.QApplication(sys.argv)
    win = Application()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
